"""
CoffeeMonitor Colombia - Motor de Exportación
===============================================
Genera archivos Excel y CSV con los datos del sistema.
"""

import csv
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

from config.settings import (
    EXPORTS_DIR,
    EXPORT_EXCEL_SHEET_MARKET,
    EXPORT_EXCEL_SHEET_COOPS,
    EXPORT_EXCEL_SHEET_ALERTS,
)
from repositories.market_repo import market_repo
from repositories.price_repo import price_repo
from repositories.alerts_repo import alerts_repo

logger = logging.getLogger(__name__)

# Estilos Excel
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
COP_FORMAT = '#,##0'
USD_FORMAT = '#,##0.00'
PCT_FORMAT = '0.0%'


class ExportService:
    """Genera exportaciones en Excel y CSV."""

    def export_market_excel(self, fecha_inicio: str, fecha_fin: str,
                            filename: str = None) -> Path:
        """Exporta datos de mercado a Excel."""
        if not filename:
            filename = f"mercado_{fecha_inicio}_{fecha_fin}.xlsx"
        filepath = EXPORTS_DIR / filename

        data = market_repo.get_range(fecha_inicio, fecha_fin)

        wb = Workbook()
        ws = wb.active
        ws.title = EXPORT_EXCEL_SHEET_MARKET

        # Título
        ws.merge_cells("A1:F1")
        ws["A1"] = f"CoffeeMonitor Colombia - Datos de Mercado ({fecha_inicio} a {fecha_fin})"
        ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2E4057")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(name="Calibri", size=9, italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        # Encabezados
        headers = ["Fecha", "Precio FNC (COP/carga)", "Bolsa NY (¢/lb)",
                    "TRM (COP/USD)", "Fuente", "Capturado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        # Datos
        for row_idx, record in enumerate(data, 5):
            ws.cell(row=row_idx, column=1, value=record["fecha"]).border = THIN_BORDER
            
            c_fnc = ws.cell(row=row_idx, column=2, value=record.get("precio_fnc"))
            c_fnc.number_format = COP_FORMAT
            c_fnc.border = THIN_BORDER
            
            c_ny = ws.cell(row=row_idx, column=3, value=record.get("bolsa_ny"))
            c_ny.number_format = USD_FORMAT
            c_ny.border = THIN_BORDER
            
            c_trm = ws.cell(row=row_idx, column=4, value=record.get("trm"))
            c_trm.number_format = COP_FORMAT
            c_trm.border = THIN_BORDER
            
            ws.cell(row=row_idx, column=5, value=record.get("fuente", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=6, value=record.get("captured_at", "")).border = THIN_BORDER

        # Ajustar anchos
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 20

        wb.save(filepath)
        logger.info(f"Exportación Excel generada: {filepath}")
        return filepath

    def export_ranking_excel(self, fecha: str, filename: str = None) -> Path:
        """Exporta ranking de cooperativas a Excel."""
        if not filename:
            filename = f"ranking_{fecha}.xlsx"
        filepath = EXPORTS_DIR / filename

        ranking = price_repo.get_ranking(fecha)

        wb = Workbook()
        ws = wb.active
        ws.title = EXPORT_EXCEL_SHEET_COOPS

        ws.merge_cells("A1:F1")
        ws["A1"] = f"CoffeeMonitor Colombia - Ranking Cooperativas ({fecha})"
        ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2E4057")

        headers = ["#", "Cooperativa", "Departamento", "Precio (COP/carga)",
                    "Factor", "Diferencial vs FNC"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

        for row_idx, record in enumerate(ranking, 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=record.get("cooperativa", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=record.get("departamento", "")).border = THIN_BORDER
            
            c = ws.cell(row=row_idx, column=4, value=record.get("precio_carga"))
            c.number_format = COP_FORMAT
            c.border = THIN_BORDER
            
            ws.cell(row=row_idx, column=5, value=record.get("factor")).border = THIN_BORDER
            
            c_dif = ws.cell(row=row_idx, column=6, value=record.get("diferencial"))
            c_dif.number_format = COP_FORMAT
            c_dif.border = THIN_BORDER
            # Resaltar diferenciales altos
            if record.get("diferencial") and record["diferencial"] > 35000:
                c_dif.font = Font(color="CC0000", bold=True)

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 22

        wb.save(filepath)
        logger.info(f"Ranking Excel generado: {filepath}")
        return filepath

    def export_alerts_excel(self, filename: str = None) -> Path:
        """Exporta alertas activas a Excel."""
        if not filename:
            filename = f"alertas_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        filepath = EXPORTS_DIR / filename

        alerts = alerts_repo.get_active(limit=200)

        wb = Workbook()
        ws = wb.active
        ws.title = EXPORT_EXCEL_SHEET_ALERTS

        headers = ["ID", "Fecha", "Regla", "Severidad", "Fuente", "Mensaje",
                    "Valor Actual", "Valor Referencia"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        severity_colors = {
            "critical": "FF0000",
            "warning": "FF8C00",
            "info": "4169E1",
        }

        for row_idx, alert in enumerate(alerts, 2):
            ws.cell(row=row_idx, column=1, value=alert["id"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=alert["fecha"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=alert["regla"]).border = THIN_BORDER
            
            sev_cell = ws.cell(row=row_idx, column=4, value=alert["severidad"])
            sev_cell.font = Font(
                color=severity_colors.get(alert["severidad"], "000000"), bold=True
            )
            sev_cell.border = THIN_BORDER
            
            ws.cell(row=row_idx, column=5, value=alert.get("source_nombre", "")).border = THIN_BORDER
            ws.cell(row=row_idx, column=6, value=alert["mensaje"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=7, value=alert.get("valor_actual")).border = THIN_BORDER
            ws.cell(row=row_idx, column=8, value=alert.get("valor_referencia")).border = THIN_BORDER

        ws.column_dimensions["F"].width = 60

        wb.save(filepath)
        logger.info(f"Alertas Excel generado: {filepath}")
        return filepath

    def export_market_csv(self, fecha_inicio: str, fecha_fin: str,
                          filename: str = None) -> Path:
        """Exporta datos de mercado a CSV."""
        if not filename:
            filename = f"mercado_{fecha_inicio}_{fecha_fin}.csv"
        filepath = EXPORTS_DIR / filename

        data = market_repo.get_range(fecha_inicio, fecha_fin)

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["fecha", "precio_fnc", "bolsa_ny", "trm", "fuente"],
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(data)

        logger.info(f"CSV generado: {filepath}")
        return filepath


export_service = ExportService()
